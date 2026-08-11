"""Exportar un reporte .xlsx de resultados: tuberías, nodos y resumen, con
las violaciones resaltadas en rojo. Una restricción desactivada simplemente no
aparece como columna — no se marca en rojo algo que no se está evaluando."""

import pandas as pd
from openpyxl.styles import PatternFill

ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _buscar_restriccion(restricciones, nombre_clase):
    return next((r for r in restricciones if type(r).__name__ == nombre_clase), None)


def _construir_tabla_tuberias(wn_final, resultados, restricciones):
    pipe_names = wn_final.pipe_name_list
    velocidad_r = _buscar_restriccion(restricciones, "VelocidadConstraint")
    hl_r = _buscar_restriccion(restricciones, "PerdidaUnitariaConstraint")

    v = resultados.link["velocity"][pipe_names].iloc[0].abs()
    hl = resultados.link["headloss"][pipe_names].iloc[0].abs() * 1000.0

    filas = []
    for name in pipe_names:
        pipe = wn_final.get_link(name)
        fila = {
            "Tuberia": name,
            "Longitud (m)": round(pipe.length, 2),
            "Diametro (mm)": round(pipe.diameter * 1000.0, 1),
            "Velocidad (m/s)": round(float(v[name]), 4),
            "Perdida unitaria (m/km)": round(float(hl[name]), 4),
        }
        if velocidad_r is not None:
            vmax_ef = velocidad_r.excepciones_vmax.get(name, velocidad_r.vmax)
            fila["Cumple velocidad"] = velocidad_r.vmin <= v[name] <= vmax_ef
        if hl_r is not None:
            fila["Cumple perdida unitaria"] = hl[name] <= hl_r.hlmax
        filas.append(fila)

    return pd.DataFrame(filas)


def _construir_tabla_nodos(wn_final, resultados, restricciones):
    junctions = wn_final.junction_name_list
    presion_r = _buscar_restriccion(restricciones, "PresionMinimaConstraint")
    presiones = resultados.node["pressure"][junctions].iloc[0]

    filas = []
    for name in junctions:
        node = wn_final.get_node(name)
        fila = {
            "Nodo": name,
            "Elevacion (m)": round(node.elevation, 2),
            "Presion (m)": round(float(presiones[name]), 3),
        }
        if presion_r is not None:
            fila["Cumple presion"] = presiones[name] >= presion_r.valor
        filas.append(fila)

    return pd.DataFrame(filas)


def _construir_tabla_resumen(wn_final, resultados, restricciones, fitness_final, catalogo):
    filas = [{"Metrica": "Fitness final", "Valor": fitness_final}]
    for r in restricciones:
        filas.append(
            {"Metrica": f"Penalizacion {type(r).__name__}", "Valor": r.evaluar(wn_final, resultados)}
        )
    filas.append({"Metrica": "Catalogo de diametros (mm)", "Valor": str(catalogo)})
    filas.append({"Metrica": "Numero de tuberias", "Valor": wn_final.num_pipes})
    return pd.DataFrame(filas)


def _resaltar_incumplimientos(ruta_salida):
    import openpyxl

    wb = openpyxl.load_workbook(ruta_salida)
    for hoja in ("Tuberias", "Nodos"):
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        encabezados = [c.value for c in ws[1]]
        columnas_cumple = [i for i, h in enumerate(encabezados) if h and h.startswith("Cumple")]
        for fila in ws.iter_rows(min_row=2):
            if any(fila[i].value is False for i in columnas_cumple):
                for celda in fila:
                    celda.fill = ROJO
    wb.save(ruta_salida)


def escribir_reporte_excel(wn_final, resultados, restricciones, ruta_salida, fitness_final, catalogo):
    df_tuberias = _construir_tabla_tuberias(wn_final, resultados, restricciones)
    df_nodos = _construir_tabla_nodos(wn_final, resultados, restricciones)
    df_resumen = _construir_tabla_resumen(wn_final, resultados, restricciones, fitness_final, catalogo)

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_tuberias.to_excel(writer, sheet_name="Tuberias", index=False)
        df_nodos.to_excel(writer, sheet_name="Nodos", index=False)
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)

    _resaltar_incumplimientos(ruta_salida)
    return ruta_salida
